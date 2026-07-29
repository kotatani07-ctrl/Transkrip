"""
Generator Excel transkrip menggunakan openpyxl.
Mengisi template yang sudah diupload berdasarkan cell_mapping.
"""
import copy
import io
import logging
from decimal import Decimal
from typing import Optional

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, column_index_from_string

from .matching import HasilPencocokan, HasilMK

logger = logging.getLogger(__name__)

# Warna highlight sesuai PRD §9
FILL_KUNING = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')   # retake
FILL_MERAH = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')    # kosong
FILL_ORANYE = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')   # tak dikenali


def _cell(ws, ref: str):
    """Akses cell dari referensi string seperti 'B5'."""
    return ws[ref]


def _parse_col(col_str: str) -> str:
    """Normalize kolom ke huruf kapital."""
    return str(col_str).strip().upper()


def _row_ref(col: str, row: int) -> str:
    return f"{_parse_col(col)}{row}"


def isi_field_biodata(ws, cm: dict, hasil: HasilPencocokan, data_pendukung=None, default_nama_dekan=''):
    """Isi field biodata ke worksheet berdasarkan cell_mapping."""
    def set_cell(key, value):
        if key in cm and cm[key]:
            try:
                ws[cm[key]] = value
            except Exception as e:
                logger.warning(f"Gagal set cell {cm[key]} untuk '{key}': {e}")

    set_cell('nim', hasil.nim)
    set_cell('nama', hasil.nama)
    set_cell('prodi', hasil.prodi)
    set_cell('jenjang', hasil.jenjang)
    set_cell('tempat_lahir', hasil.tempat_lahir)
    set_cell('tgl_lahir', hasil.tanggal_lahir)
    set_cell('tgl_lulus', hasil.tanggal_lulus)
    set_cell('angkatan', str(hasil.angkatan) if hasil.angkatan else '')
    set_cell('total_sks', hasil.total_sks)
    set_cell('ipk', float(hasil.ipk))
    set_cell('predikat', hasil.predikat)

    # Data pendukung
    if data_pendukung:
        set_cell('judul_skripsi', data_pendukung.judul_skripsi)
        set_cell('no_ijazah', data_pendukung.no_ijazah)
        nama_dekan = data_pendukung.nama_dekan or default_nama_dekan
    else:
        set_cell('judul_skripsi', '')
        set_cell('no_ijazah', '')
        nama_dekan = default_nama_dekan

    set_cell('ttd_dekan', nama_dekan)
    set_cell('nama_dekan', nama_dekan)


def isi_tabel_mk(ws, cm: dict, hasil: HasilPencocokan):
    """
    Isi tabel mata kuliah ke worksheet.
    cell_mapping harus punya 'tabel_mulai_baris' dan 'tabel_kolom' (dict kolom).
    """
    baris_mulai = cm.get('tabel_mulai_baris', 10)
    kolom = cm.get('tabel_kolom', {})

    col_no = kolom.get('no', 'A')
    col_kode = kolom.get('kode', 'B')
    col_nama = kolom.get('nama', 'C')
    col_sks = kolom.get('sks', 'D')
    col_nilai = kolom.get('nilai', 'E')
    col_lambang = kolom.get('lambang', 'F')
    col_mutu = kolom.get('mutu', 'G')

    for i, mk in enumerate(hasil.daftar_mk):
        baris = baris_mulai + i
        nomor = i + 1

        try:
            ws[_row_ref(col_no, baris)] = nomor
            ws[_row_ref(col_kode, baris)] = mk.kode_mk
            ws[_row_ref(col_nama, baris)] = mk.nama_mk
            ws[_row_ref(col_sks, baris)] = mk.sks
        except Exception as e:
            logger.warning(f"Gagal isi cell baris {baris}: {e}")
            continue

        if mk.is_kosong:
            # Warnai MERAH — tidak ada nilai
            for col in [col_no, col_kode, col_nama, col_sks]:
                try:
                    ws[_row_ref(col, baris)].fill = FILL_MERAH
                except Exception:
                    pass
        else:
            try:
                ws[_row_ref(col_nilai, baris)] = mk.nilai_angka or ''
                ws[_row_ref(col_lambang, baris)] = mk.nilai_lambang or ''
                ws[_row_ref(col_mutu, baris)] = float(mk.mutu) if mk.mutu else ''
            except Exception as e:
                logger.warning(f"Gagal isi nilai baris {baris}: {e}")

            if mk.is_retake:
                # Warnai KUNING — retake
                for col in [col_no, col_kode, col_nama, col_sks, col_nilai, col_lambang, col_mutu]:
                    try:
                        ws[_row_ref(col, baris)].fill = FILL_KUNING
                    except Exception:
                        pass


def buat_sheet_laporan(wb: openpyxl.Workbook, hasil: HasilPencocokan, sheet_name: str = 'Laporan'):
    """
    Buat sheet laporan berisi baris bermasalah + warna + keterangan.
    """
    # Hapus sheet laporan yang sudah ada dengan nama sama
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)
    headers = ['NIM', 'Nama', 'Kode MK', 'Nama MK', 'SKS', 'Nilai', 'Lambang', 'Status', 'Keterangan']
    ws.append(headers)

    # Style header
    from openpyxl.styles import Font, Alignment
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    def tambah_baris(kode, nama, sks, nilai, lambang, status, keterangan, fill_color=None):
        row_data = [hasil.nim, hasil.nama, kode, nama, sks, nilai, lambang, status, keterangan]
        ws.append(row_data)
        if fill_color:
            for cell in ws[ws.max_row]:
                cell.fill = fill_color

    # MK kosong → MERAH
    for mk in hasil.daftar_mk:
        if mk.is_kosong:
            tambah_baris(
                mk.kode_mk, mk.nama_mk, mk.sks, '', '', 'KOSONG',
                'Belum ada nilai untuk MK ini, perlu verifikasi',
                FILL_MERAH
            )

    # MK retake → KUNING
    for mk in hasil.daftar_mk:
        if mk.is_retake:
            semua = ', '.join(str(v) for v in mk.semua_nilai)
            tambah_baris(
                mk.kode_mk, mk.nama_mk, mk.sks, mk.nilai_angka or '', mk.nilai_lambang or '',
                'RETAKE',
                f'MK diulang, sistem mengambil nilai/mutu tertinggi. Semua nilai: {semua}',
                FILL_KUNING
            )

    # MK tidak dikenali → ORANYE
    for mk in hasil.mk_tidak_dikenali:
        tambah_baris(
            mk.kode_mk, mk.nama_mk, mk.sks_api, mk.nilai_angka, mk.nilai_lambang,
            'TIDAK DIKENALI',
            'Kode MK tidak dikenali, kemungkinan typo/transfer',
            FILL_ORANYE
        )

    # Auto-width kolom
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    return ws


def generate_single_excel(
    hasil: HasilPencocokan,
    template_file_path: str,
    cell_mapping: dict,
    data_pendukung=None,
    default_nama_dekan: str = '',
) -> bytes:
    """
    Generate file Excel transkrip single mahasiswa.
    Return bytes file .xlsx.
    """
    wb = load_workbook(template_file_path)
    ws = wb.active

    cm = cell_mapping
    isi_field_biodata(ws, cm, hasil, data_pendukung, default_nama_dekan)
    isi_tabel_mk(ws, cm, hasil)

    # Sheet laporan
    buat_sheet_laporan(wb, hasil, sheet_name='Laporan')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_batch_excel(
    daftar_hasil: list[tuple],  # list of (HasilPencocokan, data_pendukung, template_file_path, cell_mapping)
    default_nama_dekan: str = '',
) -> bytes:
    """
    Generate file Excel batch, dikelompokkan per (jenjang, angkatan).
    Tiap mahasiswa = blok 2 halaman dengan page break.
    Sheet Laporan gabungan di akhir.
    Return bytes .xlsx.
    """
    from openpyxl.styles import Font
    from openpyxl.worksheet.pagebreak import Break

    # Group per (jenjang, angkatan)
    from collections import defaultdict
    groups: dict = defaultdict(list)
    for item in daftar_hasil:
        hasil, data_pendukung, template_path, cm = item
        key = (hasil.jenjang, hasil.angkatan)
        groups[key].append(item)

    wb = openpyxl.Workbook()
    # Hapus sheet default
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Build sheet per grup
    for (jenjang, angkatan), items in sorted(groups.items()):
        sheet_name = f"{jenjang}_{angkatan}"[:31]  # max 31 char
        ws = wb.create_sheet(sheet_name)

        # Untuk batch: kita tulis data mahasiswa secara sederhana (tabel per mahasiswa)
        # dengan page break antar mahasiswa
        baris_offset = 1

        for idx, (hasil, data_pendukung, template_path, cm) in enumerate(items):
            # Tulis header info mahasiswa
            ws.cell(baris_offset, 1, f"NIM: {hasil.nim}")
            ws.cell(baris_offset, 1).font = Font(bold=True)
            ws.cell(baris_offset + 1, 1, f"Nama: {hasil.nama}")
            ws.cell(baris_offset + 2, 1, f"Prodi: {hasil.prodi}")
            ws.cell(baris_offset + 3, 1, f"Jenjang/Angkatan: {hasil.jenjang} / {hasil.angkatan}")
            ws.cell(baris_offset + 4, 1, f"Tgl Lulus: {hasil.tanggal_lulus}")
            ws.cell(baris_offset + 5, 1, f"IPK: {hasil.ipk}  |  Total SKS: {hasil.total_sks}")
            ws.cell(baris_offset + 6, 1, f"Predikat: {hasil.predikat}")

            if data_pendukung:
                ws.cell(baris_offset + 7, 1, f"No. Ijazah: {data_pendukung.no_ijazah}")
                ws.cell(baris_offset + 8, 1, f"Judul: {data_pendukung.judul_skripsi}")
                baris_data = baris_offset + 10
            else:
                baris_data = baris_offset + 9

            # Header tabel MK
            headers_mk = ['No', 'Kode MK', 'Nama Mata Kuliah', 'SKS', 'Nilai', 'Lambang', 'Mutu']
            for ci, h in enumerate(headers_mk, 1):
                cell = ws.cell(baris_data, ci, h)
                cell.font = Font(bold=True)

            baris_data += 1
            for no, mk in enumerate(hasil.daftar_mk, 1):
                ws.cell(baris_data, 1, no)
                ws.cell(baris_data, 2, mk.kode_mk)
                ws.cell(baris_data, 3, mk.nama_mk)
                ws.cell(baris_data, 4, mk.sks)
                if mk.is_kosong:
                    ws.cell(baris_data, 5, '')
                    ws.cell(baris_data, 6, '')
                    ws.cell(baris_data, 7, '')
                    for ci in range(1, 8):
                        ws.cell(baris_data, ci).fill = FILL_MERAH
                else:
                    ws.cell(baris_data, 5, mk.nilai_angka or '')
                    ws.cell(baris_data, 6, mk.nilai_lambang or '')
                    ws.cell(baris_data, 7, float(mk.mutu) if mk.mutu else '')
                    if mk.is_retake:
                        for ci in range(1, 8):
                            ws.cell(baris_data, ci).fill = FILL_KUNING
                baris_data += 1

            # Total
            ws.cell(baris_data, 3, 'TOTAL')
            ws.cell(baris_data, 3).font = Font(bold=True)
            ws.cell(baris_data, 4, hasil.total_sks)
            ws.cell(baris_data, 7, float(hasil.total_mutu))

            baris_akhir_mahasiswa = baris_data + 2
            baris_offset = baris_akhir_mahasiswa + 1

            # Page break setelah setiap mahasiswa (kecuali yang terakhir)
            if idx < len(items) - 1:
                ws.row_breaks.append(Break(id=baris_akhir_mahasiswa))

    # Sheet Laporan gabungan
    ws_lap = wb.create_sheet('Laporan Gabungan')
    ws_lap.append(['NIM', 'Nama', 'Kode MK', 'Nama MK', 'SKS', 'Nilai', 'Lambang', 'Status', 'Keterangan'])
    from openpyxl.styles import Font as F
    for cell in ws_lap[1]:
        cell.font = F(bold=True)

    for hasil, data_pendukung, _, _ in daftar_hasil:
        for mk in hasil.daftar_mk:
            if mk.is_kosong:
                ws_lap.append([
                    hasil.nim, hasil.nama, mk.kode_mk, mk.nama_mk, mk.sks,
                    '', '', 'KOSONG', 'Belum ada nilai untuk MK ini, perlu verifikasi'
                ])
                for cell in ws_lap[ws_lap.max_row]:
                    cell.fill = FILL_MERAH
            elif mk.is_retake:
                semua = ', '.join(str(v) for v in mk.semua_nilai)
                ws_lap.append([
                    hasil.nim, hasil.nama, mk.kode_mk, mk.nama_mk, mk.sks,
                    mk.nilai_angka or '', mk.nilai_lambang or '',
                    'RETAKE', f'Nilai diulang, diambil tertinggi. Semua: {semua}'
                ])
                for cell in ws_lap[ws_lap.max_row]:
                    cell.fill = FILL_KUNING

        for mk in hasil.mk_tidak_dikenali:
            ws_lap.append([
                hasil.nim, hasil.nama, mk.kode_mk, mk.nama_mk, mk.sks_api,
                mk.nilai_angka, mk.nilai_lambang,
                'TIDAK DIKENALI', 'Kode MK tidak dikenali, kemungkinan typo/transfer'
            ])
            for cell in ws_lap[ws_lap.max_row]:
                cell.fill = FILL_ORANYE

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
